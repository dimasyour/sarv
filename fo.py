# -*- coding: utf-8 -*-
import json

import pandas as pd
from colorama import Fore
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from tqdm import tqdm

from directory import cellAlph

URL_FO = 'http://indicators.miccedu.ru/monitoring/2019/_vpo/material.php?type=1&id=4'
ID = 'an'
HEADERS = {
    'user-agent': 'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                  'Chrome/81.0.4044.129 Safari/537.36 OPR/68.0.3618.63',
    'accept': '*/*'}


def parse_toExcel():
    print(Fore.GREEN + 'Парсинг Html-страниц в Excel...')
    with open('src/cityAndRegions.json', encoding="utf-8") as json_file:
        data = json.load(json_file)
    intIdAll = []
    arrIdCity = []
    listFileExcel = []
    linkList = []
    for x in data:
        link = str(data[x]['link'])
        intIdAll.append(int(link.split('id=')[1]))
        linkList.append(str(data[x]['link']))
    for i in range(len(intIdAll)):
        if intIdAll[i] / 10000 < 1:
            arrIdCity.append(intIdAll[i])
    with tqdm(total=len(linkList)) as pbar:
        for r in range(len(linkList)):
            url = linkList[r]
            table = pd.read_html(url, decimal=',', thousands='.')[6]
            table.to_excel('src/excel/fo/fo' + '_' + str(url.split('id=')[1]) + '.xlsx', encoding='cp1251')
            listFileExcel.append('src/excel/fo/fo' + '_' + str(url.split('id=')[1]) + '.xlsx')
            formatCell('src/excel/fo/fo' + '_' + str(url.split('id=')[1]) + '.xlsx')
            pbar.update(1)
    return listFileExcel


def formatCell(filename):
    wb1 = load_workbook(filename)
    sheet_ranges = wb1['Sheet1']
    column_c = sheet_ranges['C']
    column_d = sheet_ranges['D']
    column_e = sheet_ranges['E']
    column_f = sheet_ranges['F']
    column_g = sheet_ranges['G']
    column_h = sheet_ranges['H']
    column_i = sheet_ranges['I']
    column_j = sheet_ranges['J']
    column_k = sheet_ranges['K']
    column_l = sheet_ranges['L']
    wrapAndWidth(filename)


def wrapAndWidth(filename):
    # ширина ячейки = 60
    wb2 = load_workbook(filename)
    sheet_ranges3 = wb2['Sheet1']
    for i in range(len(cellAlph)):
        sheet_ranges3.column_dimensions[cellAlph[i]].width = 60
    # авто-перенос по словам
    for rows in sheet_ranges3.iter_rows(min_row=1, min_col=1):
        for cell in rows:
            if cell.value is not None:
                cell.alignment = Alignment(wrap_text=True)
    # удалить, всё что идет до названия города (г.)
    wb2.save(filename)
    wb4 = load_workbook(filename)
    sheet_ranges4 = wb4['Sheet1']
    column_c4 = sheet_ranges4['C']
    for i in range(len(column_c4)):
        if column_c4[i].value == '' and column_c4[i + 1].value == '':
            pass
        elif column_c4[i].value is not None:
            s = column_c4[i].value
            s1 = s.split('г.')[0]
            s2 = s1.split(' в ')[0]
            column_c4[i].value = s2
    stylesCell(sheet_ranges4)
    wb4.save(filename)


def stylesCell(sheet):
    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    sheet['C1'].fill = redFill
    sheet['B1'].value = '    '
    sheet['B2'].value = '    '
    sheet['C1'].value = '    '
    sheet['C2'].value = '    '
    # sheet.unmerge_cells(start_row=1, start_column=4, end_row=1, end_column=12)
    # sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    # sheet['A1'].value = 'Заголовок'

parse_toExcel()
