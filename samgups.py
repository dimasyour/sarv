import time
import json
import logging
from datetime import datetime
from openpyxl import load_workbook
from tqdm import tqdm

logging.basicConfig(filename="vkbot.log", level=logging.INFO)
logging.info("Excel: " + str(datetime.now()))


# данные о универе по региону (фо)
# def poRegions():
#     listK = []
#     with open('src/json/main.json', encoding="utf-8") as json_file:
#         dataVuzs = json.load(json_file)
#     listVusz = list(dataVuzs.keys())
#     for x in listVusz:
#         listK.append(dataVuzs[x])
#     return listK


def tableToTable():
    listK = []
    with open('src/json/main.json', encoding="utf-8") as json_file:
        dataVuzs = json.load(json_file)
    listVusz = list(dataVuzs.keys())
    for x in listVusz:
        listK.append(dataVuzs[x])
    wb = load_workbook(filename='src/excel.xlsx')
    ws = wb.worksheets[0]
    with tqdm(total=len(listK)) as pbar:
        for i in range(1, len(listK)):
            try:
                wb_e = load_workbook(filename='src/excel/vuzs/e/vuz_' + listK[i]['vuz_id'] + '.xlsx')
                ws_e = wb_e.worksheets[0]

                e1_1 = ws_e.cell(row=3, column=4).value
                e1_2 = ws_e.cell(row=3, column=5).value
                e1_3 = ws_e.cell(row=3, column=6).value
                e2_1 = ws_e.cell(row=4, column=4).value
                e2_2 = ws_e.cell(row=4, column=5).value
                e2_3 = ws_e.cell(row=4, column=6).value
                e3_1 = ws_e.cell(row=5, column=4).value
                e3_2 = ws_e.cell(row=5, column=5).value
                e3_3 = ws_e.cell(row=5, column=6).value
                e4_1 = ws_e.cell(row=6, column=4).value
                e4_2 = ws_e.cell(row=6, column=5).value
                e4_3 = ws_e.cell(row=6, column=6).value
                e5_1 = ws_e.cell(row=7, column=4).value
                e5_2 = ws_e.cell(row=7, column=5).value
                e5_3 = ws_e.cell(row=7, column=6).value
                e6_1 = ws_e.cell(row=8, column=4).value
                e6_2 = ws_e.cell(row=8, column=5).value
                e6_3 = ws_e.cell(row=8, column=6).value

                wb_od = load_workbook(filename='src/excel/vuzs/od/vuz_' + listK[i]['vuz_id'] + '.xlsx')
                ws_od = wb_od.worksheets[0]

                od1_1 = ws_od.cell(row=3, column=5).value
                od1_2 = ws_od.cell(row=4, column=5).value
                od1_3 = ws_od.cell(row=5, column=5).value
                od1_4 = ws_od.cell(row=6, column=5).value
                od1_5 = ws_od.cell(row=7, column=5).value
                od1_6 = ws_od.cell(row=8, column=5).value
                od1_7 = ws_od.cell(row=9, column=5).value
                od1_8 = ws_od.cell(row=10, column=5).value
                od1_9 = ws_od.cell(row=11, column=5).value
                od1_10 = ws_od.cell(row=12, column=5).value
                od1_11 = ws_od.cell(row=13, column=5).value
                od1_12 = ws_od.cell(row=14, column=5).value
                od1_13 = ws_od.cell(row=15, column=5).value
                od1_14 = ws_od.cell(row=16, column=5).value
                od1_15 = ws_od.cell(row=17, column=5).value

                wb_md = load_workbook(filename='src/excel/vuzs/md/vuz_' + listK[i]['vuz_id'] + '.xlsx')
                ws_md = wb_md.worksheets[0]

                md1_1 = ws_md.cell(row=3, column=5).value
                md1_2 = ws_md.cell(row=4, column=5).value
                md1_3 = ws_md.cell(row=5, column=5).value
                md1_4 = ws_md.cell(row=6, column=5).value
                md1_5 = ws_md.cell(row=7, column=5).value
                md1_6 = ws_md.cell(row=8, column=5).value
                md1_7 = ws_md.cell(row=9, column=5).value
                md1_8 = ws_md.cell(row=10, column=5).value
                md1_9 = ws_md.cell(row=11, column=5).value
                md1_10 = ws_md.cell(row=12, column=5).value
                md1_11 = ws_md.cell(row=13, column=5).value
                md1_12 = ws_md.cell(row=14, column=5).value
                md1_13 = ws_md.cell(row=15, column=5).value

                wb_nd = load_workbook(filename='src/excel/vuzs/nd/vuz_' + listK[i]['vuz_id'] + '.xlsx')
                ws_nd = wb_nd.worksheets[0]

                nd1_1 = ws_nd.cell(row=3, column=5).value
                nd1_2 = ws_nd.cell(row=4, column=5).value
                nd1_3 = ws_nd.cell(row=5, column=5).value
                nd1_4 = ws_nd.cell(row=6, column=5).value
                nd1_5 = ws_nd.cell(row=7, column=5).value
                nd1_6 = ws_nd.cell(row=8, column=5).value
                nd1_7 = ws_nd.cell(row=9, column=5).value
                nd1_8 = ws_nd.cell(row=10, column=5).value
                nd1_9 = ws_nd.cell(row=11, column=5).value
                nd1_10 = ws_nd.cell(row=12, column=5).value
                nd1_11 = ws_nd.cell(row=13, column=5).value
                nd1_12 = ws_nd.cell(row=14, column=5).value
                nd1_13 = ws_nd.cell(row=15, column=5).value
                nd1_14 = ws_nd.cell(row=16, column=5).value
                nd1_15 = ws_nd.cell(row=17, column=5).value
                nd1_16 = ws_nd.cell(row=18, column=5).value

                wb_fe = load_workbook(filename='src/excel/vuzs/fe/vuz_' + listK[i]['vuz_id'] + '.xlsx')
                ws_fe = wb_fe.worksheets[0]

                fe1_1 = ws_fe.cell(row=3, column=5).value
                fe1_2 = ws_fe.cell(row=4, column=5).value
                fe1_3 = ws_fe.cell(row=5, column=5).value
                fe1_4 = ws_fe.cell(row=6, column=5).value

                wb_inf = load_workbook(filename='src/excel/vuzs/inf/vuz_' + listK[i]['vuz_id'] + '.xlsx')
                ws_inf = wb_inf.worksheets[0]

                inf1_1 = ws_inf.cell(row=3, column=5).value
                inf1_2 = ws_inf.cell(row=4, column=5).value
                inf1_3 = ws_inf.cell(row=5, column=5).value
                inf1_4 = ws_inf.cell(row=6, column=5).value
                inf1_5 = ws_inf.cell(row=7, column=5).value
                inf1_6 = ws_inf.cell(row=8, column=5).value
                inf1_7 = ws_inf.cell(row=9, column=5).value
                inf1_8 = ws_inf.cell(row=10, column=5).value

                wb_kadr = load_workbook(filename='src/excel/vuzs/kadr/vuz_' + listK[i]['vuz_id'] + '.xlsx')
                ws_kadr = wb_kadr.worksheets[0]

                kadr1_1 = ws_kadr.cell(row=3, column=5).value
                kadr1_2 = ws_kadr.cell(row=4, column=5).value
                kadr1_3 = ws_kadr.cell(row=5, column=5).value
                kadr1_4 = ws_kadr.cell(row=6, column=5).value
                kadr1_5 = ws_kadr.cell(row=7, column=5).value

                wb_forRegion = load_workbook(filename='src/excel/vuzs/forRegion/vuz_' + listK[i]['vuz_id'] + '.xlsx')
                ws_forRegion = wb_forRegion.worksheets[0]

                forRegion1_1 = ws_forRegion.cell(row=2, column=2).value
                forRegion1_2 = ws_forRegion.cell(row=3, column=2).value

                wb_allInRegion = load_workbook(filename='src/excel/vuzs/allInRegion/vuz_' + listK[i]['vuz_id'] + '.xlsx')
                ws_allInRegion = wb_allInRegion.worksheets[0]

                allInRegion1_1 = ws_allInRegion.cell(row=2, column=2).value

                wb_studentInVuz = load_workbook(filename='src/excel/vuzs/studentInVuz/vuz_' + listK[i]['vuz_id'] + '.xlsx')
                ws_studentInVuz = wb_studentInVuz.worksheets[0]

                studentInVuz1_1 = ws_studentInVuz.cell(row=2, column=2).value
                studentInVuz1_2 = ws_studentInVuz.cell(row=3, column=2).value

                wb_procentStudent = load_workbook(
                    filename='src/excel/vuzs/procentStudent/vuz_' + listK[i]['vuz_id'] + '.xlsx')
                ws_procentStudent = wb_procentStudent.worksheets[0]

                procentStudent1_1 = ws_procentStudent.cell(row=2, column=3).value
                procentStudent1_2 = ws_procentStudent.cell(row=2, column=5).value

                wb_dolyaInRegion = load_workbook(
                    filename='src/excel/vuzs/dolyaInRegion/vuz_' + listK[i]['vuz_id'] + '.xlsx')
                ws_dolyaInRegion = wb_dolyaInRegion.worksheets[0]

                dolyaInRegion1_1 = ws_dolyaInRegion.cell(row=2, column=3).value
                dolyaInRegion1_2 = ws_dolyaInRegion.cell(row=2, column=4).value
                dolyaInRegion1_3 = ws_dolyaInRegion.cell(row=3, column=3).value
                dolyaInRegion1_4 = ws_dolyaInRegion.cell(row=3, column=4).value
                dolyaInRegion1_5 = ws_dolyaInRegion.cell(row=4, column=3).value
                dolyaInRegion1_6 = ws_dolyaInRegion.cell(row=4, column=4).value
                dolyaInRegion1_7 = ws_dolyaInRegion.cell(row=5, column=3).value
                dolyaInRegion1_8 = ws_dolyaInRegion.cell(row=5, column=4).value
                dolyaInRegion1_9 = ws_dolyaInRegion.cell(row=6, column=3).value
                dolyaInRegion1_10 = ws_dolyaInRegion.cell(row=6, column=4).value
                dolyaInRegion1_11 = ws_dolyaInRegion.cell(row=7, column=3).value
                dolyaInRegion1_12 = ws_dolyaInRegion.cell(row=7, column=4).value
                dolyaInRegion1_13 = ws_dolyaInRegion.cell(row=8, column=3).value
                dolyaInRegion1_14 = ws_dolyaInRegion.cell(row=8, column=4).value
                dolyaInRegion1_15 = ws_dolyaInRegion.cell(row=9, column=3).value
                dolyaInRegion1_16 = ws_dolyaInRegion.cell(row=9, column=4).value

                wb_dop = load_workbook(filename='src/excel/vuzs/dop/vuz_' + listK[i]['vuz_id'] + '.xlsx')
                ws_dop = wb_dop.worksheets[0]

                dop1_1 = ws_dop.cell(row=5, column=5).value
                dop1_2 = ws_dop.cell(row=6, column=5).value
                dop1_3 = ws_dop.cell(row=7, column=5).value
                dop1_4 = ws_dop.cell(row=8, column=5).value
                dop1_5 = ws_dop.cell(row=9, column=5).value
                dop1_6 = ws_dop.cell(row=10, column=5).value
                dop1_7 = ws_dop.cell(row=11, column=5).value
                dop1_8 = ws_dop.cell(row=12, column=5).value
                dop1_9 = ws_dop.cell(row=13, column=5).value
                dop1_10 = ws_dop.cell(row=14, column=5).value
                dop1_11 = ws_dop.cell(row=15, column=5).value
                dop1_12 = ws_dop.cell(row=16, column=5).value

                dop2_1 = ws_dop.cell(row=18, column=5).value
                dop2_2 = ws_dop.cell(row=19, column=5).value
                dop2_3 = ws_dop.cell(row=20, column=5).value
                dop2_4 = ws_dop.cell(row=21, column=5).value
                dop2_5 = ws_dop.cell(row=22, column=5).value
                dop2_6 = ws_dop.cell(row=23, column=5).value
                dop2_7 = ws_dop.cell(row=24, column=5).value
                dop2_8 = ws_dop.cell(row=25, column=5).value
                dop2_9 = ws_dop.cell(row=26, column=5).value
                dop2_10 = ws_dop.cell(row=27, column=5).value
                dop2_11 = ws_dop.cell(row=28, column=5).value

                dop3_1 = ws_dop.cell(row=30, column=5).value
                dop3_2 = ws_dop.cell(row=31, column=5).value
                dop3_3 = ws_dop.cell(row=32, column=5).value
                dop3_4 = ws_dop.cell(row=33, column=5).value
                dop3_5 = ws_dop.cell(row=34, column=5).value
                dop3_6 = ws_dop.cell(row=35, column=5).value
                dop3_7 = ws_dop.cell(row=36, column=5).value
                dop3_8 = ws_dop.cell(row=37, column=5).value
                dop3_9 = ws_dop.cell(row=38, column=5).value

                dop4_1 = ws_dop.cell(row=40, column=5).value
                dop4_2 = ws_dop.cell(row=41, column=5).value
                dop4_3 = ws_dop.cell(row=42, column=5).value
                dop4_4 = ws_dop.cell(row=43, column=5).value
                dop4_5 = ws_dop.cell(row=44, column=5).value
                dop4_6 = ws_dop.cell(row=45, column=5).value
                dop4_7 = ws_dop.cell(row=46, column=5).value
                dop4_8 = ws_dop.cell(row=47, column=5).value
                dop4_9 = ws_dop.cell(row=48, column=5).value

                dop5_1 = ws_dop.cell(row=50, column=5).value
                dop5_2 = ws_dop.cell(row=51, column=5).value
                dop5_3 = ws_dop.cell(row=52, column=5).value
                dop5_4 = ws_dop.cell(row=53, column=5).value
                dop5_5 = ws_dop.cell(row=54, column=5).value
                dop5_6 = ws_dop.cell(row=55, column=5).value
                dop5_7 = ws_dop.cell(row=56, column=5).value
                dop5_8 = ws_dop.cell(row=57, column=5).value
                dop5_9 = ws_dop.cell(row=58, column=5).value

                dop6_1 = ws_dop.cell(row=60, column=5).value
                dop6_2 = ws_dop.cell(row=61, column=5).value
                dop6_3 = ws_dop.cell(row=62, column=5).value
                dop6_4 = ws_dop.cell(row=63, column=5).value
                dop6_5 = ws_dop.cell(row=64, column=5).value
                dop6_6 = ws_dop.cell(row=65, column=5).value
                dop6_7 = ws_dop.cell(row=66, column=5).value
                dop6_8 = ws_dop.cell(row=67, column=5).value
                dop6_9 = ws_dop.cell(row=68, column=5).value

                ws.cell(row=i, column=1).value = listK[i]['vuz_name']
                ws.cell(row=i, column=2).value = listK[i]['vuz_link']

                ws.cell(row=i, column=3).value = e1_1
                ws.cell(row=i, column=4).value = e1_2
                ws.cell(row=i, column=5).value = e1_3

                ws.cell(row=i, column=6).value = e2_1
                ws.cell(row=i, column=7).value = e2_2
                ws.cell(row=i, column=8).value = e2_3

                ws.cell(row=i, column=9).value = e3_1
                ws.cell(row=i, column=10).value = e3_2
                ws.cell(row=i, column=11).value = e3_3

                ws.cell(row=i, column=11).value = e4_1
                ws.cell(row=i, column=12).value = e4_2
                ws.cell(row=i, column=13).value = e4_3

                ws.cell(row=i, column=14).value = e5_1
                ws.cell(row=i, column=15).value = e5_2
                ws.cell(row=i, column=16).value = e5_3

                ws.cell(row=i, column=17).value = e6_1
                ws.cell(row=i, column=18).value = e6_2
                ws.cell(row=i, column=19).value = e6_3

                ws.cell(row=i, column=20).value = od1_1
                ws.cell(row=i, column=21).value = od1_2
                ws.cell(row=i, column=22).value = od1_3
                ws.cell(row=i, column=23).value = od1_4
                ws.cell(row=i, column=24).value = od1_5
                ws.cell(row=i, column=25).value = od1_6
                ws.cell(row=i, column=26).value = od1_7
                ws.cell(row=i, column=27).value = od1_8
                ws.cell(row=i, column=28).value = od1_9
                ws.cell(row=i, column=29).value = od1_10
                ws.cell(row=i, column=30).value = od1_11
                ws.cell(row=i, column=31).value = od1_12
                ws.cell(row=i, column=32).value = od1_13
                ws.cell(row=i, column=33).value = od1_14
                ws.cell(row=i, column=34).value = od1_15

                ws.cell(row=i, column=35).value = md1_1
                ws.cell(row=i, column=36).value = md1_2
                ws.cell(row=i, column=37).value = md1_3
                ws.cell(row=i, column=38).value = md1_4
                ws.cell(row=i, column=39).value = md1_5
                ws.cell(row=i, column=40).value = md1_6
                ws.cell(row=i, column=41).value = md1_7
                ws.cell(row=i, column=42).value = md1_8
                ws.cell(row=i, column=43).value = md1_9
                ws.cell(row=i, column=44).value = md1_10
                ws.cell(row=i, column=45).value = md1_11
                ws.cell(row=i, column=46).value = md1_12
                ws.cell(row=i, column=47).value = md1_13

                ws.cell(row=i, column=48).value = nd1_1
                ws.cell(row=i, column=49).value = nd1_2
                ws.cell(row=i, column=50).value = nd1_3
                ws.cell(row=i, column=51).value = nd1_4
                ws.cell(row=i, column=52).value = nd1_5
                ws.cell(row=i, column=53).value = nd1_6
                ws.cell(row=i, column=54).value = nd1_7
                ws.cell(row=i, column=55).value = nd1_8
                ws.cell(row=i, column=56).value = nd1_9
                ws.cell(row=i, column=57).value = nd1_10
                ws.cell(row=i, column=58).value = nd1_11
                ws.cell(row=i, column=59).value = nd1_12
                ws.cell(row=i, column=60).value = nd1_13
                ws.cell(row=i, column=61).value = nd1_14
                ws.cell(row=i, column=62).value = nd1_15
                ws.cell(row=i, column=63).value = nd1_16

                ws.cell(row=i, column=64).value = fe1_1
                ws.cell(row=i, column=65).value = fe1_2
                ws.cell(row=i, column=66).value = fe1_3
                ws.cell(row=i, column=67).value = fe1_4

                ws.cell(row=i, column=68).value = inf1_1
                ws.cell(row=i, column=69).value = inf1_2
                ws.cell(row=i, column=70).value = inf1_3
                ws.cell(row=i, column=71).value = inf1_4
                ws.cell(row=i, column=72).value = inf1_5
                ws.cell(row=i, column=73).value = inf1_6
                ws.cell(row=i, column=74).value = inf1_7
                ws.cell(row=i, column=75).value = inf1_8

                ws.cell(row=i, column=76).value = kadr1_1
                ws.cell(row=i, column=77).value = kadr1_2
                ws.cell(row=i, column=78).value = kadr1_3
                ws.cell(row=i, column=79).value = kadr1_4
                ws.cell(row=i, column=80).value = kadr1_5

                ws.cell(row=i, column=81).value = forRegion1_1
                ws.cell(row=i, column=82).value = forRegion1_2

                ws.cell(row=i, column=83).value = allInRegion1_1

                ws.cell(row=i, column=84).value = studentInVuz1_1
                ws.cell(row=i, column=85).value = studentInVuz1_2

                ws.cell(row=i, column=86).value = procentStudent1_1
                ws.cell(row=i, column=87).value = procentStudent1_2

                ws.cell(row=i, column=88).value = dolyaInRegion1_1
                ws.cell(row=i, column=89).value = dolyaInRegion1_2
                ws.cell(row=i, column=90).value = dolyaInRegion1_3
                ws.cell(row=i, column=91).value = dolyaInRegion1_4
                ws.cell(row=i, column=92).value = dolyaInRegion1_5
                ws.cell(row=i, column=93).value = dolyaInRegion1_6
                ws.cell(row=i, column=94).value = dolyaInRegion1_7
                ws.cell(row=i, column=95).value = dolyaInRegion1_8
                ws.cell(row=i, column=96).value = dolyaInRegion1_9
                ws.cell(row=i, column=97).value = dolyaInRegion1_10
                ws.cell(row=i, column=98).value = dolyaInRegion1_11
                ws.cell(row=i, column=99).value = dolyaInRegion1_12
                ws.cell(row=i, column=100).value = dolyaInRegion1_13
                ws.cell(row=i, column=101).value = dolyaInRegion1_14
                ws.cell(row=i, column=102).value = dolyaInRegion1_15
                ws.cell(row=i, column=103).value = dolyaInRegion1_16

                ws.cell(row=i, column=104).value = dop1_1
                ws.cell(row=i, column=105).value = dop1_2
                ws.cell(row=i, column=106).value = dop1_3
                ws.cell(row=i, column=107).value = dop1_4
                ws.cell(row=i, column=108).value = dop1_5
                ws.cell(row=i, column=109).value = dop1_6
                ws.cell(row=i, column=110).value = dop1_7
                ws.cell(row=i, column=111).value = dop1_8
                ws.cell(row=i, column=112).value = dop1_9
                ws.cell(row=i, column=113).value = dop1_10
                ws.cell(row=i, column=114).value = dop1_11
                ws.cell(row=i, column=115).value = dop1_12

                ws.cell(row=i, column=116).value = dop2_1
                ws.cell(row=i, column=117).value = dop2_2
                ws.cell(row=i, column=118).value = dop2_3
                ws.cell(row=i, column=119).value = dop2_4
                ws.cell(row=i, column=120).value = dop2_5
                ws.cell(row=i, column=121).value = dop2_6
                ws.cell(row=i, column=122).value = dop2_7
                ws.cell(row=i, column=123).value = dop2_8
                ws.cell(row=i, column=124).value = dop2_9
                ws.cell(row=i, column=125).value = dop2_10
                ws.cell(row=i, column=126).value = dop2_11

                ws.cell(row=i, column=127).value = dop3_1
                ws.cell(row=i, column=128).value = dop3_2
                ws.cell(row=i, column=129).value = dop3_3
                ws.cell(row=i, column=130).value = dop3_4
                ws.cell(row=i, column=131).value = dop3_5
                ws.cell(row=i, column=132).value = dop3_6
                ws.cell(row=i, column=133).value = dop3_7
                ws.cell(row=i, column=134).value = dop3_8
                ws.cell(row=i, column=135).value = dop3_9

                ws.cell(row=i, column=136).value = dop4_1
                ws.cell(row=i, column=137).value = dop4_2
                ws.cell(row=i, column=138).value = dop4_3
                ws.cell(row=i, column=139).value = dop4_4
                ws.cell(row=i, column=140).value = dop4_5
                ws.cell(row=i, column=141).value = dop4_6
                ws.cell(row=i, column=142).value = dop4_7
                ws.cell(row=i, column=143).value = dop4_8
                ws.cell(row=i, column=144).value = dop4_9

                ws.cell(row=i, column=145).value = dop5_1
                ws.cell(row=i, column=146).value = dop5_2
                ws.cell(row=i, column=147).value = dop5_3
                ws.cell(row=i, column=148).value = dop5_4
                ws.cell(row=i, column=149).value = dop5_5
                ws.cell(row=i, column=150).value = dop5_6
                ws.cell(row=i, column=151).value = dop5_7
                ws.cell(row=i, column=152).value = dop5_8
                ws.cell(row=i, column=153).value = dop5_9

                ws.cell(row=i, column=154).value = dop6_1
                ws.cell(row=i, column=155).value = dop6_2
                ws.cell(row=i, column=156).value = dop6_3
                ws.cell(row=i, column=157).value = dop6_4
                ws.cell(row=i, column=158).value = dop6_5
                ws.cell(row=i, column=159).value = dop6_6
                ws.cell(row=i, column=160).value = dop6_7
                ws.cell(row=i, column=161).value = dop6_8
                ws.cell(row=i, column=162).value = dop6_9

                ws.cell(row=i, column=163).value = dataVuzs[i]['fo']
                ws.cell(row=i, column=164).value = dataVuzs[i]['region']
                pbar.update(1)

            except Exception as e:
                logging.error(str(datetime.now()) + " " + str(e) + str(" : ") + str(listK[i]['vuz_id']))
                pbar.update(1) 
    wb.save('src/excel.xlsx')


tableToTable()
# print(poRegions())
